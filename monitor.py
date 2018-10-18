import os
import lcm
import sys
import time
import subprocess
from Sim import SimHandler
from PyQt5.QtQuick import QQuickView
from PyQt5.QtGui import QGuiApplication
from PyQt5.QtCore import QUrl, QObject, pyqtSignal, QTimer, QThread, pyqtSlot
import openpyxl


wb = openpyxl.load_workbook('./files/scenario.xlsx')
ws = wb.worksheets[0]
rows = ws.max_row + 1
dataOne = []
dataTwo = []
dataThree = []
dataFour = []


class MainPanel(QObject):
    def __init__(self, parent=None):
        super(QObject, self).__init__(parent=parent)
        # 初始化一个定时器
        self.timer = QTimer(self)
        # 定义时间超时连接start_app
        self.timer.timeout.connect(self.start)
        # 定义时间任务是一次性任务
        self.timer.setSingleShot(False)
        # 启动时间任务
        self.timer.start()
        # 实例化一个线程
        self.work = WorkThread()
        # 多线程的信号触发连接到up_data
        self.work.trigger.connect(self.up_data)

    def start(self):
        self.work.start()

    @pyqtSlot()
    def startWork(self):
        subprocess.Popen('.\simulator\Simulator.exe geely', shell=True)

    @pyqtSlot()
    def stopWork(self):
        lc = lcm.LCM("udpm://239.255.76.67:7667?ttl=1")
        msg = SimHandler()
        msg.timestamp = int(time.time() * 1000000)
        msg.parameters = str("")
        msg.handler = -2
        lc.publish("SimulatorCtrl", msg.encode())

    @pyqtSlot()
    def dataRead(self):
        self.readExcelData()

    def up_data(self):
        rootObject.setProperty("handler", handler)

    def readExcelData(self):
        for row in range(2, rows):
            dataOne.append(ws['B' + str(row)].value)
            dataTwo.append(ws['C' + str(row)].value)
            dataThree.append(ws['D' + str(row)].value)
            dataFour.append(ws['E' + str(row)].value)
        rootObject.setProperty("dataOne", dataOne)
        rootObject.setProperty("dataTwo", dataTwo)
        rootObject.setProperty("dataThree", dataThree)
        rootObject.setProperty("dataFour", dataFour)


channelName = ""
timestamp = ""
handler = ""
parameters = ""


class HandleWork(MainPanel):
    def my_handler(channel, data):
        msg = SimHandler.decode(data)
        global channelName
        global timestamp
        global handler
        global parameters

        channelName = channel
        timestamp = str(msg.timestamp)
        handler = str(msg.handler)
        parameters = str(msg.parameters)

    lc = lcm.LCM()
    subscription = lc.subscribe("SimulatorCtrl", my_handler)


class WorkThread(QThread):
    # 定义一个信号
    trigger = pyqtSignal(str)

    def __int__(self):
        # 初始化函数，默认
        super(WorkThread, self).__init__()

    def run(self):
        # time.sleep(2)
        self.trigger.emit("ok")
        HandleWork.lc.handle()


if __name__ == "__main__":
    path = './content/MainControl.qml'
    app = QGuiApplication(sys.argv)
    viewer = QQuickView()
    con = MainPanel()
    context = viewer.rootContext()
    context.setContextProperty("con", con)  
    viewer.engine().quit.connect(app.quit)
    viewer.setResizeMode(QQuickView.SizeRootObjectToView)
    viewer.setSource(QUrl(path))
    rootObject = viewer.rootObject()
    viewer.show()
    app.exec_()
